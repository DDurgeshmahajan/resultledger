"""
Utility functions for Excel formatting and branding.
Designed by Durgesh Mahajan
"""

import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def brand_excel(filepath: str) -> None:
    """
    Apply professional branding and formatting to the output Excel file.
    Adds a header row with branding, styles all columns, and auto-sizes.
    """
    wb = load_workbook(filepath)
    ws = wb.active

    # --- Insert branding rows at the top ---
    ws.insert_rows(1, 3)

    # Row 1: Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = "Savitribai Phule Pune University — Result Ledger Extract"
    title_cell.font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="1A237E", end_color="1A237E", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Row 2: Subtitle / Credit
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ws.max_column)
    subtitle_cell = ws.cell(row=2, column=1)
    subtitle_cell.value = "Extraction Tool "
    subtitle_cell.font = Font(name="Calibri", size=11, italic=True, color="FFFFFF")
    subtitle_cell.fill = PatternFill(start_color="283593", end_color="283593", fill_type="solid")
    subtitle_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Row 3: Spacer (empty)
    for col in range(1, ws.max_column + 1):
        ws.cell(row=3, column=col).fill = PatternFill(
            start_color="E8EAF6", end_color="E8EAF6", fill_type="solid"
        )

    # --- Style the header row (now row 4) ---
    header_row = 4
    header_fill = PatternFill(start_color="3949AB", end_color="3949AB", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin", color="B0BEC5"),
        right=Side(style="thin", color="B0BEC5"),
        top=Side(style="thin", color="B0BEC5"),
        bottom=Side(style="thin", color="B0BEC5"),
    )

    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    # --- Style data rows ---
    data_font = Font(name="Calibri", size=10)
    alt_fill_1 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    alt_fill_2 = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

    for row in range(header_row + 1, ws.max_row + 1):
        fill = alt_fill_1 if (row - header_row) % 2 == 1 else alt_fill_2
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = data_font
            cell.fill = fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # --- Auto-size columns ---
    for col in range(1, ws.max_column + 1):
        max_len = 10  # minimum width
        col_letter = get_column_letter(col)
        for row in range(header_row, ws.max_row + 1):
            val = ws.cell(row=row, column=col).value
            if val:
                max_len = max(max_len, len(str(val)) + 2)
        ws.column_dimensions[col_letter].width = min(max_len, 35)

    # Set row heights
    ws.row_dimensions[1].height = 35
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 8
    ws.row_dimensions[header_row].height = 28

    # Freeze panes below the header
    ws.freeze_panes = f"A{header_row + 1}"

    wb.save(filepath)


def get_output_filename(original_filename: str) -> str:
    """Generate output Excel filename from the original PDF filename."""
    base = os.path.splitext(original_filename)[0]
    # Clean the filename
    base = base.replace(" ", "_")
    return f"{base}.xlsx"
